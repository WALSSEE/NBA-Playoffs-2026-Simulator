import streamlit as st
import pandas as pd
import numpy as np
import io

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="NBA Playoff Odds",
    page_icon="🏀",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Custom CSS ────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Bebas+Neue&family=Inter:wght@300;400;500;600&display=swap');
html, body, [class*="css"] { font-family: 'Inter', sans-serif; }
h1, h2, h3 { font-family: 'Bebas Neue', sans-serif; letter-spacing: 2px; }
.stApp { background: #0a0a0f; color: #e8e8e8; }
[data-testid="stSidebar"] { background: #111118 !important; border-right: 1px solid #222230; }
.metric-card {
    background: linear-gradient(135deg, #1a1a2e 0%, #16213e 100%);
    border: 1px solid #2a2a4a; border-radius: 12px;
    padding: 20px; text-align: center; margin-bottom: 10px;
}
.metric-value {
    font-family: 'Bebas Neue', sans-serif; font-size: 2.8rem;
    color: #f97316; line-height: 1; margin: 0;
}
.metric-label { font-size: 0.75rem; color: #8888aa; text-transform: uppercase; letter-spacing: 1.5px; margin-top: 4px; }
hr { border-color: #222230 !important; }
.stButton > button {
    background: linear-gradient(135deg, #f97316, #ea580c) !important;
    color: white !important; border: none !important;
    font-family: 'Bebas Neue', sans-serif !important; font-size: 1.1rem !important;
    letter-spacing: 2px !important; border-radius: 8px !important;
    padding: 8px 24px !important;
}
.stTabs [data-baseweb="tab-list"] { background: #111118; border-bottom: 2px solid #f97316; gap: 4px; }
.stTabs [data-baseweb="tab"] { font-family: 'Bebas Neue', sans-serif; letter-spacing: 1.5px; font-size: 1rem; color: #8888aa !important; }
.stTabs [aria-selected="true"] { color: #f97316 !important; background: #1a1a2e !important; }
.playin-box {
    background: linear-gradient(135deg, #1a0a2e 0%, #0a1a2e 100%);
    border: 1px solid #7c3aed; border-radius: 10px; padding: 14px 18px; margin: 8px 0;
}
</style>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════
# MATH HELPERS
# ══════════════════════════════════════════════════════════

def win_prob(power_a: float, power_b: float, home_advantage: float = 3.0) -> float:
    """Logistic win probability. home_advantage added to power_a's side."""
    diff = (power_a + home_advantage) - power_b
    return float(np.clip(1 / (1 + np.exp(-diff * 0.155)), 0.001, 0.999))


def series_probs(p_home: float, p_away: float, best_of: int = 7) -> dict:
    """
    Exact series probabilities (dynamic programming).
    p_home = P(team A wins) when A has home court.
    p_away = P(team A wins) when B has home court.
    NBA 2-2-1-1-1 schedule.
    """
    wins_needed = (best_of + 1) // 2
    game_is_home = {1:True,2:True,3:False,4:False,5:True,6:False,7:True}

    states = {(0, 0): 1.0}
    outcome_probs = {}

    for g in range(1, best_of + 1):
        pa = p_home if game_is_home[g] else p_away
        new_states = {}
        for (wa, wb), prob in states.items():
            for da, db, pw in [(1, 0, pa), (0, 1, 1 - pa)]:
                nwa, nwb = wa + da, wb + db
                if nwa == wins_needed or nwb == wins_needed:
                    outcome_probs[(nwa, nwb)] = outcome_probs.get((nwa, nwb), 0) + prob * pw
                else:
                    new_states[(nwa, nwb)] = new_states.get((nwa, nwb), 0) + prob * pw
        states = new_states

    total_a = sum(v for (wa, wb), v in outcome_probs.items() if wa == wins_needed)
    total_b = sum(v for (wa, wb), v in outcome_probs.items() if wb == wins_needed)
    a_wins  = {k: v for k, v in outcome_probs.items() if k[0] == wins_needed}
    b_wins  = {k: v for k, v in outcome_probs.items() if k[1] == wins_needed}

    def mgn(d, m):
        return sum(v for (wa, wb), v in d.items() if abs(wa - wb) >= m)

    return {
        'p_a': total_a, 'p_b': total_b, 'outcomes': outcome_probs,
        'p_a_margin_3': mgn(a_wins, 3), 'p_a_margin_2': mgn(a_wins, 2), 'p_a_margin_1': mgn(a_wins, 1),
        'p_b_margin_3': mgn(b_wins, 3), 'p_b_margin_2': mgn(b_wins, 2), 'p_b_margin_1': mgn(b_wins, 1),
    }


def playin_probs_exact(t7, t8, t9, t10, power_map, home_adv):
    """
    Analytical Play-In probabilities.
    Game 1: t7 vs t8 at t7's court.
    Game 2: t9 vs t10 at t9's court.
    Game 3: loser(G1) vs winner(G2) at loser(G1)'s court (always better seed).
    Returns dict: team -> {p7, p8, p_qualify}
    """
    p78  = win_prob(power_map[t7], power_map[t8],  home_adv)   # P(t7 wins G1)
    p910 = win_prob(power_map[t9], power_map[t10], home_adv)   # P(t9 wins G2)

    # G3 outcomes depend on who lost G1 and who won G2
    p_t8_vs_t9  = win_prob(power_map[t8], power_map[t9],  home_adv)  # t8 hosts (seed 8 < 9)
    p_t8_vs_t10 = win_prob(power_map[t8], power_map[t10], home_adv)
    p_t7_vs_t9  = win_prob(power_map[t7], power_map[t9],  home_adv)  # t7 hosts (seed 7 < 9/10)
    p_t7_vs_t10 = win_prob(power_map[t7], power_map[t10], home_adv)

    res = {t: {'p7': 0.0, 'p8': 0.0} for t in [t7, t8, t9, t10]}

    # Branch A: t7 wins G1 (prob=p78) → t7 qualifies as seed 7, t8 plays G3
    res[t7]['p7'] += p78
    res[t8]['p8'] += p78 * p910      * p_t8_vs_t9      # t8 beats t9 in G3
    res[t9]['p8'] += p78 * p910      * (1-p_t8_vs_t9)  # t9 beats t8 in G3
    res[t8]['p8'] += p78 * (1-p910)  * p_t8_vs_t10
    res[t10]['p8']+= p78 * (1-p910)  * (1-p_t8_vs_t10)

    # Branch B: t8 wins G1 (prob=1-p78) → t8 qualifies as seed 7, t7 plays G3
    res[t8]['p7'] += (1-p78)
    res[t7]['p8'] += (1-p78) * p910      * p_t7_vs_t9
    res[t9]['p8'] += (1-p78) * p910      * (1-p_t7_vs_t9)
    res[t7]['p8'] += (1-p78) * (1-p910)  * p_t7_vs_t10
    res[t10]['p8']+= (1-p78) * (1-p910)  * (1-p_t7_vs_t10)

    for t in res:
        res[t]['p_qualify'] = res[t]['p7'] + res[t]['p8']
    return res


def sim_series_mc(ta, tb, power_map, seed_map, home_adv, rng, best_of=7):
    """Monte Carlo single series simulation."""
    wn = (best_of + 1) // 2
    home = ta if seed_map[ta] < seed_map[tb] else tb
    away = tb if home == ta else ta
    hp, ap = power_map[home], power_map[away]
    # NBA 2-2-1-1-1: True = home team is at their court
    schedule = [True, True, False, False, True, False, True]
    wa = wb = 0
    for i in range(best_of):
        if schedule[i]:
            # Home court game for 'home' team
            p_home_wins = win_prob(hp, ap, home_adv)
            home_wins = rng.random() < p_home_wins
        else:
            # Away team's home court
            p_away_wins = win_prob(ap, hp, home_adv)
            home_wins = not (rng.random() < p_away_wins)
        winner = home if home_wins else away
        if winner == ta: wa += 1
        else: wb += 1
        if wa == wn: return ta
        if wb == wn: return tb
    return ta


def sim_playin_mc(by_seed, power_map, home_adv, rng):
    """Simulate Play-In once. Returns (seed7_qualifier, seed8_qualifier)."""
    t7, t8  = by_seed[7], by_seed[8]
    t9, t10 = by_seed[9], by_seed[10]

    # G1: t7 vs t8, t7 at home
    p_g1 = win_prob(power_map[t7], power_map[t8], home_adv)
    g1_winner = t7 if rng.random() < p_g1 else t8
    g1_loser  = t8 if g1_winner == t7 else t7

    # G2: t9 vs t10, t9 at home
    p_g2 = win_prob(power_map[t9], power_map[t10], home_adv)
    g2_winner = t9 if rng.random() < p_g2 else t10

    # G3: g1_loser vs g2_winner, g1_loser at home (better seed)
    p_g3 = win_prob(power_map[g1_loser], power_map[g2_winner], home_adv)
    g3_winner = g1_loser if rng.random() < p_g3 else g2_winner

    return g1_winner, g3_winner  # (7th seed, 8th seed)


def sim_conf_with_playin(teams_10, home_adv, n_sim, rng):
    """Full conference simulation including Play-In."""
    power_map = {t['name']: t['power'] for t in teams_10}
    seed_map  = {t['name']: t['seed']  for t in teams_10}
    by_seed   = {t['seed']: t['name']  for t in teams_10}

    conf_wins = {t['name']: 0 for t in teams_10}

    for _ in range(n_sim):
        # Play-In → get seeds 7 and 8
        q7, q8 = sim_playin_mc(by_seed, power_map, home_adv, rng)

        # Playoff bracket: seeds 1-6 direct + q7, q8
        bracket = [by_seed[s] for s in range(1, 7)] + [q7, q8]
        bseed   = {bracket[i]: i+1 for i in range(8)}

        def sim_s(ta, tb):
            return sim_series_mc(ta, tb, power_map, bseed, home_adv, rng)

        # R1: 1v8, 2v7, 3v6, 4v5
        r1 = [sim_s(bracket[0], bracket[7]),
              sim_s(bracket[1], bracket[6]),
              sim_s(bracket[2], bracket[5]),
              sim_s(bracket[3], bracket[4])]

        # R2: W(1/8) vs W(4/5), W(2/7) vs W(3/6)
        r2 = [sim_s(r1[0], r1[3]),
              sim_s(r1[1], r1[2])]

        # Conf Final
        cf = sim_s(r2[0], r2[1])
        conf_wins[cf] += 1

    return conf_wins


def fmt_pct(v: float) -> str:
    return f"{v*100:.1f}%"


# ══════════════════════════════════════════════════════════
# SIDEBAR
# ══════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("## ⚙️ Asetukset")
    home_adv = st.slider("Kotietu (pistettä)", 0.0, 6.0, 3.0, 0.5,
        help="Kotijoukkueen powerratin korotus yhdessä ottelussa")
    n_sim = st.select_slider("Simulaatioiden määrä",
        options=[10_000, 50_000, 100_000, 200_000], value=100_000,
        format_func=lambda x: f"{x:,}")
    st.divider()
    st.markdown("### 📁 Tuo Excelistä")
    uploaded = st.file_uploader("Lataa joukkueiden voimaluvut (.xlsx / .csv)",
        type=["xlsx","csv"])
    df_up = None
    if uploaded:
        try:
            df_up = pd.read_csv(uploaded) if uploaded.name.endswith(".csv") else pd.read_excel(uploaded)
            st.success(f"Ladattu {len(df_up)} joukkuetta")
            st.dataframe(df_up, use_container_width=True, height=200)
        except Exception as e:
            st.error(f"Virhe: {e}")
    st.divider()
    st.markdown("### 🔮 Play-In rakenne")
    st.markdown("""
**Peli 1:** #7 vs #8 → voittaja = sija 7  
**Peli 2:** #9 vs #10 → häviäjä putoaa  
**Peli 3:** häviäjä(P1) vs voittaja(P2) → voittaja = sija 8  
Kotietu paremmalle siemenelle joka pelissä.
""")

# ══════════════════════════════════════════════════════════
# HEADER
# ══════════════════════════════════════════════════════════
st.markdown("""
<div style="background:linear-gradient(135deg,#0f0f1a,#1a0a2e,#0f1a0a);
     border:1px solid #f97316; border-radius:12px; padding:24px 32px;
     margin-bottom:24px; text-align:center;">
    <h1 style="font-size:3rem; margin:0; color:#f97316;">🏀 NBA PLAYOFF ODDS</h1>
    <p style="color:#8888aa; margin:8px 0 0; font-size:0.9rem; letter-spacing:2px;">
        PUDOTUSPELISARJOJEN TODENNÄKÖISYYSLASKURI · PLAY-IN TURNAUS MUKANA
    </p>
</div>
""", unsafe_allow_html=True)

tab1, tab2, tab3, tab4 = st.tabs(["🎯 SARJA-ANALYYSI", "🔮 PLAY-IN", "🏆 KONFERENSSI & MESTARUUS", "📊 EXCEL-POHJA"])


# ════════════════════════════════════════════════════════
# TAB 1 – Single series analysis
# ════════════════════════════════════════════════════════
with tab1:
    st.markdown("### Syötä joukkueiden tiedot")
    c1, c2 = st.columns(2)
    with c1:
        st.markdown('<div style="color:#f97316;font-family:Bebas Neue,sans-serif;font-size:1.2rem;letter-spacing:2px;">🟠 KORKEAMPI SIJOITUS (KOTIETU)</div>', unsafe_allow_html=True)
        name_a  = st.text_input("Joukkue A", value="Boston Celtics", key="na")
        power_a = st.number_input("Joukkue A voimaluku", value=115.0, step=0.5, key="pa")
    with c2:
        st.markdown('<div style="color:#3b82f6;font-family:Bebas Neue,sans-serif;font-size:1.2rem;letter-spacing:2px;">🔵 MATALAMPI SIJOITUS (VIERASJOUKKUE)</div>', unsafe_allow_html=True)
        name_b  = st.text_input("Joukkue B", value="New York Knicks", key="nb")
        power_b = st.number_input("Joukkue B voimaluku", value=112.0, step=0.5, key="pb")

    best_of = st.radio("Sarjamuoto", [5, 7], index=1, horizontal=True,
                       format_func=lambda x: f"Best-of-{x}")

    if st.button("🔢 LASKE TODENNÄKÖISYYDET", use_container_width=True):
        p_home = win_prob(power_a, power_b,  home_adv)
        p_away = win_prob(power_a, power_b, -home_adv)
        res    = series_probs(p_home, p_away, best_of)
        wins_n = (best_of + 1) // 2

        st.divider()
        st.markdown("#### Yhden pelin voittotodennäköisyys")
        mc1, mc2 = st.columns(2)
        with mc1:
            st.markdown(f'<div class="metric-card"><p class="metric-value">{p_home*100:.1f}%</p><p class="metric-label">{name_a} – kotipeli</p></div>', unsafe_allow_html=True)
        with mc2:
            st.markdown(f'<div class="metric-card"><p class="metric-value" style="color:#3b82f6;">{(1-p_away)*100:.1f}%</p><p class="metric-label">{name_b} – kotipeli</p></div>', unsafe_allow_html=True)

        st.divider()
        st.markdown("#### Sarjan voittotodennäköisyys")
        col_a, col_b = st.columns(2)
        with col_a:
            st.markdown(f'<div class="metric-card"><p class="metric-value">{res["p_a"]*100:.1f}%</p><p class="metric-label">{name_a} voittaa sarjan</p></div>', unsafe_allow_html=True)
        with col_b:
            st.markdown(f'<div class="metric-card"><p class="metric-value" style="color:#3b82f6;">{res["p_b"]*100:.1f}%</p><p class="metric-label">{name_b} voittaa sarjan</p></div>', unsafe_allow_html=True)

        bar_a, bar_b = res['p_a']*100, res['p_b']*100
        st.markdown(f"""
        <div style="display:flex;gap:4px;margin:12px 0;">
            <div style="width:{bar_a:.1f}%;background:linear-gradient(90deg,#f97316,#fb923c);
                 border-radius:6px 0 0 6px;padding:6px 10px;color:white;font-weight:700;
                 font-size:0.85rem;white-space:nowrap;overflow:hidden;">{name_a} {bar_a:.1f}%</div>
            <div style="width:{bar_b:.1f}%;background:linear-gradient(90deg,#3b82f6,#60a5fa);
                 border-radius:0 6px 6px 0;padding:6px 10px;color:white;font-weight:700;
                 font-size:0.85rem;text-align:right;white-space:nowrap;overflow:hidden;">{bar_b:.1f}% {name_b}</div>
        </div>""", unsafe_allow_html=True)

        st.divider()
        st.markdown("#### Marginaalilinjat")
        margin_rows = [{"Linja": lbl,
                        f"{name_a} kattaa": fmt_pct(res[f'p_a_margin_{m}']),
                        f"{name_b} kattaa": fmt_pct(res[f'p_b_margin_{m}'])}
                       for m, lbl in [(3,"+3.5"),(2,"+2.5"),(1,"+1.5")]]
        st.dataframe(pd.DataFrame(margin_rows).set_index("Linja"), use_container_width=True)

        st.divider()
        st.markdown("#### Sarjan lopputulokset")
        oc_rows = []
        for (wa, wb), prob in sorted(res['outcomes'].items(), key=lambda x: -x[1]):
            winner = name_a if wa == wins_n else name_b
            loser  = name_b if wa == wins_n else name_a
            score  = f"{wa}–{wb}" if wa == wins_n else f"{wb}–{wa}"
            oc_rows.append({"Tulos": f"{winner} {score} {loser}", "Voittaja": winner,
                            "Todennäköisyys": fmt_pct(prob), "Tn %": round(prob*100,2)})
        df_oc = pd.DataFrame(oc_rows)
        st.dataframe(df_oc[["Tulos","Voittaja","Todennäköisyys"]], use_container_width=True, hide_index=True)
        st.bar_chart(df_oc.set_index("Tulos")["Tn %"], color="#f97316", height=250)


# ════════════════════════════════════════════════════════
# TAB 2 – Play-In tournament
# ════════════════════════════════════════════════════════
with tab2:
    st.markdown("### Play-In turnauksen todennäköisyydet")
    st.markdown("""
<div class="playin-box">
<b style="color:#a78bfa;font-family:Bebas Neue,sans-serif;letter-spacing:2px;font-size:1.1rem;">🔮 PLAY-IN RAKENNE</b><br><br>
<b>Peli 1:</b> #7 vs #8 &nbsp;(kotietu: #7) &nbsp;→ <b>Voittaja = playoff sija #7</b><br>
<b>Peli 2:</b> #9 vs #10 (kotietu: #9) → <b>Häviäjä putoaa kaudelta</b><br>
<b>Peli 3:</b> Häviäjä(P1) vs Voittaja(P2) (kotietu: häviäjä(P1), parempi siemen) → <b>Voittaja = playoff sija #8</b>
</div>
""", unsafe_allow_html=True)

    st.markdown("#### Syötä Play-In joukkueet")
    pi_cols = st.columns(4)
    pi_names, pi_powers = [], []
    defaults_pi = [
        ("Philadelphia 76ers", 107.0), ("Miami Heat", 105.0),
        ("Chicago Bulls", 103.5),      ("Atlanta Hawks", 102.0),
    ]
    for i, (col, (dname, dpower)) in enumerate(zip(pi_cols, defaults_pi)):
        with col:
            st.markdown(f'<div style="color:#a78bfa;font-family:Bebas Neue,sans-serif;letter-spacing:1px;font-size:1rem;">SIJA #{i+7}</div>', unsafe_allow_html=True)
            n = st.text_input("Nimi", value=dname, key=f"pi_n{i}", label_visibility="collapsed")
            p = st.number_input("Voima", value=dpower, step=0.5, key=f"pi_p{i}", label_visibility="collapsed")
            pi_names.append(n); pi_powers.append(p)

    if st.button("🔮 LASKE PLAY-IN TODENNÄKÖISYYDET", use_container_width=True):
        pm = {pi_names[i]: pi_powers[i] for i in range(4)}
        t7, t8, t9, t10 = pi_names

        probs = playin_probs_exact(t7, t8, t9, t10, pm, home_adv)

        # Per-game win probs
        st.divider()
        st.markdown("#### Yksittäisten pelien todennäköisyydet")
        g1 = win_prob(pm[t7], pm[t8],  home_adv)
        g2 = win_prob(pm[t9], pm[t10], home_adv)
        g_cols = st.columns(3)
        with g_cols[0]:
            st.markdown(f"""<div class="metric-card">
                <p style="color:#a78bfa;font-family:Bebas Neue,sans-serif;letter-spacing:1px;margin:0 0 4px;">PELI 1</p>
                <p style="margin:2px 0;font-size:0.8rem;">{t7} vs {t8}</p>
                <p class="metric-value">{g1*100:.1f}%</p>
                <p class="metric-label">{t7} voittaa</p></div>""", unsafe_allow_html=True)
        with g_cols[1]:
            st.markdown(f"""<div class="metric-card">
                <p style="color:#a78bfa;font-family:Bebas Neue,sans-serif;letter-spacing:1px;margin:0 0 4px;">PELI 2</p>
                <p style="margin:2px 0;font-size:0.8rem;">{t9} vs {t10}</p>
                <p class="metric-value">{g2*100:.1f}%</p>
                <p class="metric-label">{t9} voittaa</p></div>""", unsafe_allow_html=True)
        with g_cols[2]:
            st.markdown(f"""<div class="metric-card">
                <p style="color:#a78bfa;font-family:Bebas Neue,sans-serif;letter-spacing:1px;margin:0 0 4px;">PELI 3</p>
                <p style="margin:2px 0;font-size:0.8rem;">Riippuu peleistä 1 & 2</p>
                <p class="metric-value" style="font-size:1.8rem;">4 sk.</p>
                <p class="metric-label">Katso alla</p></div>""", unsafe_allow_html=True)

        # Qualification probabilities
        st.divider()
        st.markdown("#### Pääsy playoffseihin – kokonaistodennäköisyydet")
        pi_res_rows = []
        for i, name in enumerate(pi_names):
            p = probs[name]
            pi_res_rows.append({
                "Joukkue": name, "Lähtösija": f"#{i+7}", "Voimaluku": pi_powers[i],
                "Sija #7 %": fmt_pct(p['p7']),
                "Sija #8 %": fmt_pct(p['p8']),
                "Pääsee playoffseihin %": fmt_pct(p['p_qualify']),
                "_q": p['p_qualify'],
            })
        df_pi = pd.DataFrame(pi_res_rows).sort_values("_q", ascending=False)
        st.dataframe(df_pi[["Joukkue","Lähtösija","Voimaluku","Sija #7 %","Sija #8 %","Pääsee playoffseihin %"]],
                     use_container_width=True, hide_index=True)

        # Visual bars
        st.markdown("#### Visualisointi")
        for _, row in df_pi.iterrows():
            name = row["Joukkue"]
            p7   = probs[name]['p7']  * 100
            p8   = probs[name]['p8']  * 100
            pout = (1 - probs[name]['p_qualify']) * 100
            st.markdown(f"**{name}** ({row['Lähtösija']})")
            st.markdown(f"""
            <div style="display:flex;gap:2px;margin:3px 0 10px;">
                <div style="width:{p7:.1f}%;background:linear-gradient(90deg,#f97316,#fb923c);
                     border-radius:5px 0 0 5px;padding:4px 8px;color:white;font-size:0.75rem;
                     font-weight:700;white-space:nowrap;overflow:hidden;min-width:0;">Sija #7: {p7:.1f}%</div>
                <div style="width:{p8:.1f}%;background:linear-gradient(90deg,#7c3aed,#a78bfa);
                     padding:4px 8px;color:white;font-size:0.75rem;font-weight:700;
                     white-space:nowrap;overflow:hidden;min-width:0;">Sija #8: {p8:.1f}%</div>
                <div style="width:{pout:.1f}%;background:#1a1a2e;border-radius:0 5px 5px 0;
                     padding:4px 8px;color:#8888aa;font-size:0.75rem;
                     white-space:nowrap;overflow:hidden;min-width:0;">Putoaa: {pout:.1f}%</div>
            </div>""", unsafe_allow_html=True)

        # All G3 scenarios
        st.divider()
        st.markdown("#### Peli 3 – kaikki skenaariot")
        p78v  = win_prob(pm[t7], pm[t8],  home_adv)
        p910v = win_prob(pm[t9], pm[t10], home_adv)
        scenarios = [
            (f"t7 voittaa P1, t9 voittaa P2 → t8 vs t9",  p78v     * p910v,       t8, t9 ),
            (f"t7 voittaa P1, t10 voittaa P2 → t8 vs t10", p78v    * (1-p910v),    t8, t10),
            (f"t8 voittaa P1, t9 voittaa P2 → t7 vs t9",  (1-p78v) * p910v,        t7, t9 ),
            (f"t8 voittaa P1, t10 voittaa P2 → t7 vs t10",(1-p78v) * (1-p910v),    t7, t10),
        ]
        # Replace t7/t8/t9/t10 with real names
        sc_rows = []
        for desc, prob_s, home_t, away_t in scenarios:
            desc_real = desc.replace("t7", t7).replace("t8", t8).replace("t9", t9).replace("t10", t10)
            pg3 = win_prob(pm[home_t], pm[away_t], home_adv)
            sc_rows.append({
                "Skenaario": desc_real,
                "Todennäköisyys": fmt_pct(prob_s),
                f"{home_t} voittaa G3": fmt_pct(pg3),
                f"{away_t} voittaa G3": fmt_pct(1-pg3),
            })
        st.dataframe(pd.DataFrame(sc_rows), use_container_width=True, hide_index=True)


# ════════════════════════════════════════════════════════
# TAB 3 – Full conference + NBA championship
# ════════════════════════════════════════════════════════
with tab3:
    st.markdown("### Konferenssi- ja NBA-mestaruustodennäköisyydet")
    st.caption("Syötä molempien konferenssien **10 joukkuetta** (sijat 1–10). "
               "Play-In (7–10) simuloidaan automaattisesti.")

    default_east_10 = [
        ("Boston Celtics",        1, 118.0), ("Cleveland Cavaliers",  2, 115.5),
        ("New York Knicks",       3, 113.0), ("Indiana Pacers",        4, 111.5),
        ("Milwaukee Bucks",       5, 110.0), ("Orlando Magic",         6, 108.5),
        ("Philadelphia 76ers",    7, 107.0), ("Miami Heat",            8, 105.0),
        ("Chicago Bulls",         9, 103.5), ("Atlanta Hawks",        10, 102.0),
    ]
    default_west_10 = [
        ("Oklahoma City Thunder", 1, 119.0), ("Denver Nuggets",        2, 116.0),
        ("Minnesota Timberwolves",3, 114.0), ("LA Clippers",           4, 112.5),
        ("Dallas Mavericks",      5, 111.0), ("Phoenix Suns",          6, 109.5),
        ("Sacramento Kings",      7, 108.0), ("Golden State Warriors", 8, 106.0),
        ("Memphis Grizzlies",     9, 104.5), ("New Orleans Pelicans", 10, 103.0),
    ]

    def build_inputs_10(conf_label, defaults, key_prefix):
        teams = []
        with st.expander(f"🏀 {conf_label}nen konferenssi – 10 joukkuetta", expanded=True):
            h = st.columns([3,1,2])
            h[0].markdown("**Joukkue**"); h[1].markdown("**Sija**"); h[2].markdown("**Voimaluku**")
            for i, (dname, dseed, dpower) in enumerate(defaults):
                if dseed == 7:
                    st.markdown('<div style="border-top:1px dashed #7c3aed;margin:4px 0;color:#a78bfa;font-size:0.72rem;letter-spacing:1px;padding-top:4px;">▼ PLAY-IN JOUKKUEET (7–10)</div>', unsafe_allow_html=True)
                c1, c2, c3 = st.columns([3,1,2])
                name  = c1.text_input("", value=dname,   key=f"{key_prefix}_n{i}", label_visibility="collapsed")
                seed  = c2.number_input("", value=dseed, min_value=1, max_value=10,
                                        key=f"{key_prefix}_s{i}", label_visibility="collapsed")
                power = c3.number_input("", value=dpower, step=0.5,
                                        key=f"{key_prefix}_p{i}", label_visibility="collapsed")
                teams.append({"name": name, "seed": seed, "power": power})
        return teams

    east_10 = build_inputs_10("Itäi", default_east_10, "e10")
    west_10 = build_inputs_10("Länti", default_west_10, "w10")

    if st.button("🏆 SIMULOI KOKO PLAYOFF-BRACKET", use_container_width=True):
        rng = np.random.default_rng(42)

        with st.spinner(f"Simuloidaan {n_sim:,} kautta Play-In + Playoffs + Finaalit..."):
            east_cwins = sim_conf_with_playin(sorted(east_10, key=lambda t: t['seed']), home_adv, n_sim, rng)
            west_cwins = sim_conf_with_playin(sorted(west_10, key=lambda t: t['seed']), home_adv, n_sim, rng)

        # Play-In summary
        st.divider()
        st.markdown("#### 🔮 Play-In: Todennäköisyys päästä playoffseihin")
        col_ep, col_wp = st.columns(2)
        for col, teams_10, label in [(col_ep, east_10, "Itä"), (col_wp, west_10, "Länsi")]:
            with col:
                st.markdown(f"**{label}inen Play-In**")
                bsp  = {t['seed']: t['name'] for t in teams_10}
                pmap = {t['name']: t['power'] for t in teams_10}
                pi_r = playin_probs_exact(bsp[7], bsp[8], bsp[9], bsp[10], pmap, home_adv)
                pi_rows = [{"Joukkue": bsp[s], "Sija": f"#{s}",
                            "Sija #7 %": fmt_pct(pi_r[bsp[s]]['p7']),
                            "Sija #8 %": fmt_pct(pi_r[bsp[s]]['p8']),
                            "Pääsee %":  fmt_pct(pi_r[bsp[s]]['p_qualify'])}
                           for s in [7,8,9,10]]
                st.dataframe(pd.DataFrame(pi_rows), use_container_width=True, hide_index=True)

        # Conference results
        st.divider()
        st.markdown("#### 🏆 Konferenssimestaruus")
        col_e, col_w = st.columns(2)
        for col, teams_10, cwins, label, color in [
            (col_e, east_10, east_cwins, "Itä",  "#f97316"),
            (col_w, west_10, west_cwins, "Länsi","#3b82f6"),
        ]:
            with col:
                st.markdown(f"**{label}inen konferenssi**")
                rows_c = [{"Joukkue": t['name'] + (" 🔮" if t['seed'] >= 7 else ""),
                           "Sija": t['seed'], "Voima": t['power'],
                           "Konf. Mestari %": round(cwins[t['name']]/n_sim*100, 1)}
                          for t in sorted(teams_10, key=lambda x: x['seed'])]
                df_c = pd.DataFrame(rows_c)
                st.dataframe(df_c, use_container_width=True, hide_index=True)
                st.bar_chart(df_c.set_index("Joukkue")["Konf. Mestari %"], color=color, height=200)

        # NBA Championship
        st.divider()
        st.markdown("#### 🏆 NBA Mestaruus")
        all_power = {t['name']: t['power'] for t in east_10 + west_10}
        nba_champ = {t['name']: 0.0 for t in east_10 + west_10}

        east_cp = {t['name']: east_cwins[t['name']]/n_sim for t in east_10}
        west_cp = {t['name']: west_cwins[t['name']]/n_sim for t in west_10}

        for et in east_10:
            for wt in west_10:
                pair_p = east_cp[et['name']] * west_cp[wt['name']]
                if pair_p < 1e-5:
                    continue
                pa, pb = all_power[et['name']], all_power[wt['name']]
                if pa >= pb:
                    r = series_probs(win_prob(pa, pb, home_adv), win_prob(pa, pb, -home_adv), 7)
                    nba_champ[et['name']] += pair_p * r['p_a']
                    nba_champ[wt['name']] += pair_p * r['p_b']
                else:
                    r = series_probs(win_prob(pb, pa, home_adv), win_prob(pb, pa, -home_adv), 7)
                    nba_champ[wt['name']] += pair_p * r['p_a']
                    nba_champ[et['name']] += pair_p * r['p_b']

        all_rows = []
        for t in east_10 + west_10:
            conf = "Itä" if t in east_10 else "Länsi"
            cw   = east_cwins[t['name']] if t in east_10 else west_cwins[t['name']]
            all_rows.append({
                "Joukkue": t['name'] + (" 🔮" if t['seed'] >= 7 else ""),
                "Konf.": conf, "Sija": t['seed'], "Voima": t['power'],
                "Konf. Mestari %": round(cw/n_sim*100, 1),
                "NBA Mestari %": round(nba_champ[t['name']]*100, 1),
            })

        all_df = pd.DataFrame(all_rows).sort_values("NBA Mestari %", ascending=False)
        st.dataframe(all_df[["Joukkue","Konf.","Sija","Voima","Konf. Mestari %","NBA Mestari %"]],
                     use_container_width=True, hide_index=True)
        st.bar_chart(all_df.set_index("Joukkue")["NBA Mestari %"], color="#f97316", height=340)
        st.caption("🔮 = joukkue tuli playoffseihin Play-In turnauksen kautta")


# ════════════════════════════════════════════════════════
# TAB 4 – Excel template
# ════════════════════════════════════════════════════════
with tab4:
    st.markdown("### 📊 Excel-pohja voimalukuille")
    st.markdown("Lataa valmis pohja (10 joukkuetta per konferenssi, Play-In joukkueet merkitty). "
                "Täytä voimaluvut ja lataa sivupalkista.")

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Joukkueet"

    headers = ["Joukkue", "Konferenssi", "Sija", "Voima"]
    hfill = PatternFill("solid", fgColor="1a1a2e")
    hfont = Font(bold=True, color="F97316", name="Calibri", size=12)
    thin  = Side(border_style="thin", color="2a2a4a")
    brd   = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(1, ci, h)
        cell.font, cell.fill = hfont, hfill
        cell.alignment, cell.border = Alignment(horizontal="center"), brd

    sample20 = [
        ("Boston Celtics","Itä",1,118.0),("Cleveland Cavaliers","Itä",2,115.5),
        ("New York Knicks","Itä",3,113.0),("Indiana Pacers","Itä",4,111.5),
        ("Milwaukee Bucks","Itä",5,110.0),("Orlando Magic","Itä",6,108.5),
        ("Philadelphia 76ers","Itä",7,107.0),("Miami Heat","Itä",8,105.0),
        ("Chicago Bulls","Itä",9,103.5),("Atlanta Hawks","Itä",10,102.0),
        ("Oklahoma City Thunder","Länsi",1,119.0),("Denver Nuggets","Länsi",2,116.0),
        ("Minnesota Timberwolves","Länsi",3,114.0),("LA Clippers","Länsi",4,112.5),
        ("Dallas Mavericks","Länsi",5,111.0),("Phoenix Suns","Länsi",6,109.5),
        ("Sacramento Kings","Länsi",7,108.0),("Golden State Warriors","Länsi",8,106.0),
        ("Memphis Grizzlies","Länsi",9,104.5),("New Orleans Pelicans","Länsi",10,103.0),
    ]

    dfont  = Font(name="Calibri", size=11, color="E8E8E8")
    dfill  = PatternFill("solid", fgColor="0a0a0f")
    afill  = PatternFill("solid", fgColor="111118")
    pifill = PatternFill("solid", fgColor="1a0a2e")  # Play-In highlight

    for ri, row in enumerate(sample20, 2):
        is_playin = row[2] >= 7
        fill = pifill if is_playin else (dfill if ri % 2 == 0 else afill)
        for ci, val in enumerate(row, 1):
            cell = ws.cell(ri, ci, val)
            cell.font, cell.fill = dfont, fill
            cell.alignment = Alignment(horizontal="center" if ci > 1 else "left")
            cell.border = brd

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 12

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)

    st.download_button("⬇️ LATAA EXCEL-POHJA (20 joukkuetta)",
        data=buf, file_name="nba_voimaluvut_playin.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True)

    st.divider()
    st.markdown("""
### 📖 Ohjeet
**Play-In joukkueet (sijat 7–10)** on korostettu violetilla taustavärillä Excel-pohjassa.

**Suositeltuja voimalukulähteitä:**
- **Net Rating** (OffRtg − DefRtg) – NBA.com, ESPN
- **Adjusted Net Rating** – Cleaning the Glass, BBall-Index
- **RAPTOR / EPM** – FiveThirtyEight, BBALL-Index

**Kotietu:** NBA:ssa historiallinen kotivoitto% ~59–62 %, vastaa ~2.5–3.5 pistettä. Oletus 3.0 on hyvä.
""")
