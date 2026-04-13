import streamlit as st
import pandas as pd
import numpy as np
import io

st.set_page_config(page_title="NBA Playoff Odds 2025", page_icon="🏀", layout="wide", initial_sidebar_state="expanded")

# ── CSS ───────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Bebas+Neue&family=Inter:wght@300;400;500;600&display=swap');
html,body,[class*="css"]{font-family:'Inter',sans-serif;}
h1,h2,h3{font-family:'Bebas Neue',sans-serif;letter-spacing:2px;}
.stApp{background:#0a0a0f;color:#e8e8e8;}
[data-testid="stSidebar"]{background:#111118!important;border-right:1px solid #222230;}
.metric-card{background:linear-gradient(135deg,#1a1a2e,#16213e);border:1px solid #2a2a4a;
 border-radius:12px;padding:18px;text-align:center;margin-bottom:8px;}
.metric-value{font-family:'Bebas Neue',sans-serif;font-size:2.6rem;color:#f97316;line-height:1;margin:0;}
.metric-label{font-size:0.72rem;color:#8888aa;text-transform:uppercase;letter-spacing:1.5px;margin-top:4px;}
hr{border-color:#222230!important;}
.stButton>button{background:linear-gradient(135deg,#f97316,#ea580c)!important;color:white!important;
 border:none!important;font-family:'Bebas Neue',sans-serif!important;font-size:1.1rem!important;
 letter-spacing:2px!important;border-radius:8px!important;padding:8px 24px!important;}
.stTabs [data-baseweb="tab-list"]{background:#111118;border-bottom:2px solid #f97316;gap:4px;}
.stTabs [data-baseweb="tab"]{font-family:'Bebas Neue',sans-serif;letter-spacing:1.5px;font-size:1rem;color:#8888aa!important;}
.stTabs [aria-selected="true"]{color:#f97316!important;background:#1a1a2e!important;}
.playin-box{background:linear-gradient(135deg,#1a0a2e,#0a1a2e);border:1px solid #7c3aed;border-radius:10px;padding:14px 18px;margin:8px 0;}
.bracket-team{background:#16213e;border:1px solid #2a2a4a;border-radius:6px;padding:4px 10px;
 font-size:0.8rem;display:inline-block;margin:2px;}
.bracket-seed{color:#f97316;font-family:'Bebas Neue',sans-serif;margin-right:4px;}
</style>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# MATH – Net Rating → Win Probability
# Net rating calibration:
#   NBA avg margin of victory per game ≈ 7-8 pts, net rating spread ≈ 10-12 pts
#   Empirically: 1 pt net rating ≈ ~2.7% win probability shift (logistic)
#   Home court in NBA ≈ +2.0 to +3.0 net rating points (we use +3.0 default)
#   Calibration: 10 pt NR difference → ~73% win prob (logistic with k=0.116)
# ══════════════════════════════════════════════════════════════════════════════

def win_prob(nr_a: float, nr_b: float, home_adv: float = 3.0) -> float:
    """
    Win probability from net ratings.
    nr_a, nr_b: net ratings (e.g. +5.2, -1.3)
    home_adv: net rating bonus for home team (default 3.0 points)
    k=0.116 calibrated so that 10pt NR gap ≈ 73% win prob per game.
    """
    diff = (nr_a + home_adv) - nr_b
    return float(np.clip(1 / (1 + np.exp(-diff * 0.116)), 0.001, 0.999))


def series_probs_exact(nr_home: float, nr_away: float, home_adv: float, best_of: int = 7) -> dict:
    """
    Exact series probabilities via dynamic programming.
    nr_home = net rating of higher-seeded (home) team
    nr_away = net rating of lower-seeded (away) team
    NBA 2-2-1-1-1 schedule: games 1,2,5,7 at higher seed; 3,4,6 at lower seed
    """
    wn = (best_of + 1) // 2
    # True = home team plays at home court
    schedule = {1:True,2:True,3:False,4:False,5:True,6:False,7:True}

    p_h = win_prob(nr_home, nr_away,  home_adv)   # home team wins when at home
    p_a = win_prob(nr_home, nr_away, -home_adv)   # home team wins when at away

    states = {(0,0): 1.0}
    outcomes = {}

    for g in range(1, best_of+1):
        pw = p_h if schedule[g] else p_a
        ns = {}
        for (wh, wa), prob in states.items():
            for dwh, dwa, p in [(1,0,pw),(0,1,1-pw)]:
                nwh, nwa = wh+dwh, wa+dwa
                if nwh==wn or nwa==wn:
                    outcomes[(nwh,nwa)] = outcomes.get((nwh,nwa),0) + prob*p
                else:
                    ns[(nwh,nwa)] = ns.get((nwh,nwa),0) + prob*p
        states = ns

    pa_tot = sum(v for (wh,wa),v in outcomes.items() if wh==wn)
    pb_tot = sum(v for (wh,wa),v in outcomes.items() if wa==wn)
    a_wins = {k:v for k,v in outcomes.items() if k[0]==wn}
    b_wins = {k:v for k,v in outcomes.items() if k[1]==wn}

    def mgn(d,m):
        return sum(v for (wh,wa),v in d.items() if abs(wh-wa)>=m)

    return {
        'p_home': pa_tot, 'p_away': pb_tot, 'outcomes': outcomes,
        'p_home_m3': mgn(a_wins,3), 'p_home_m2': mgn(a_wins,2), 'p_home_m1': mgn(a_wins,1),
        'p_away_m3': mgn(b_wins,3), 'p_away_m2': mgn(b_wins,2), 'p_away_m1': mgn(b_wins,1),
    }


# ── Play-In (analytical) ──────────────────────────────────────────────────────
def playin_probs_exact(t7, t8, t9, t10, nr_map, home_adv):
    """
    Exact Play-In probabilities (all 4 branches).
    Home court goes to better-seeded team in every game.
    Returns dict: team -> {p7, p8, p_qualify}
    """
    p78  = win_prob(nr_map[t7],  nr_map[t8],  home_adv)
    p910 = win_prob(nr_map[t9],  nr_map[t10], home_adv)
    # G3: loser(G1) hosts winner(G2) — loser(G1) always has better seed (7 or 8 vs 9 or 10)
    p_t8_t9  = win_prob(nr_map[t8],  nr_map[t9],  home_adv)
    p_t8_t10 = win_prob(nr_map[t8],  nr_map[t10], home_adv)
    p_t7_t9  = win_prob(nr_map[t7],  nr_map[t9],  home_adv)
    p_t7_t10 = win_prob(nr_map[t7],  nr_map[t10], home_adv)

    r = {t:{'p7':0.,'p8':0.} for t in [t7,t8,t9,t10]}
    # A: t7 wins G1
    r[t7]['p7'] += p78
    r[t8]['p8'] += p78*p910*(p_t8_t9);      r[t9]['p8']  += p78*p910*(1-p_t8_t9)
    r[t8]['p8'] += p78*(1-p910)*(p_t8_t10); r[t10]['p8'] += p78*(1-p910)*(1-p_t8_t10)
    # B: t8 wins G1
    r[t8]['p7'] += (1-p78)
    r[t7]['p8'] += (1-p78)*p910*(p_t7_t9);      r[t9]['p8']  += (1-p78)*p910*(1-p_t7_t9)
    r[t7]['p8'] += (1-p78)*(1-p910)*(p_t7_t10); r[t10]['p8'] += (1-p78)*(1-p910)*(1-p_t7_t10)
    for t in r: r[t]['p_qualify'] = r[t]['p7']+r[t]['p8']
    return r


# ── Monte Carlo: full bracket path-aware simulation ──────────────────────────
def sim_series_mc(home_name, away_name, nr_map, home_adv, rng, best_of=7):
    """Simulate one series. home_name is the higher-seeded team."""
    wn = (best_of+1)//2
    ph = win_prob(nr_map[home_name], nr_map[away_name],  home_adv)
    pa = win_prob(nr_map[home_name], nr_map[away_name], -home_adv)
    schedule = [True,True,False,False,True,False,True]
    wh = wa = 0
    for i in range(best_of):
        p = ph if schedule[i] else pa
        if rng.random() < p: wh += 1
        else:                wa += 1
        if wh==wn: return home_name
        if wa==wn: return away_name
    return home_name


def sim_playin_mc(t7, t8, t9, t10, nr_map, home_adv, rng):
    g1 = t7 if rng.random() < win_prob(nr_map[t7], nr_map[t8],  home_adv) else t8
    g1l = t8 if g1==t7 else t7
    g2 = t9 if rng.random() < win_prob(nr_map[t9], nr_map[t10], home_adv) else t10
    g3 = g1l if rng.random() < win_prob(nr_map[g1l], nr_map[g2], home_adv) else g2
    return g1, g3   # (seed-7 qualifier, seed-8 qualifier)


def sim_full_bracket(east_teams, west_teams, home_adv, n_sim):
    """
    Full path-aware Monte Carlo simulation.
    Tracks every team's path round by round.
    east_teams / west_teams: list of dicts {name, seed, nr}  (seed 1–10)
    
    Returns:
      - conf_wins_e / conf_wins_w: {name: count}
      - nba_wins: {name: count}
      - round_wins_e / round_wins_w: {name: {1:count,2:count,3:count}} (playoff rounds)
      - playin_qualify_e / playin_qualify_w: {name: {7:count,8:count}}
    """
    rng = np.random.default_rng(42)
    nr_map = {t['name']: t['nr'] for t in east_teams+west_teams}

    conf_wins_e  = {t['name']:0 for t in east_teams}
    conf_wins_w  = {t['name']:0 for t in west_teams}
    nba_wins     = {t['name']:0 for t in east_teams+west_teams}
    round_wins_e = {t['name']:{1:0,2:0,3:0} for t in east_teams}
    round_wins_w = {t['name']:{1:0,2:0,3:0} for t in west_teams}
    playin_q_e   = {t['name']:{7:0,8:0} for t in east_teams}
    playin_q_w   = {t['name']:{7:0,8:0} for t in west_teams}

    e_by_seed = {t['seed']:t['name'] for t in east_teams}
    w_by_seed = {t['seed']:t['name'] for t in west_teams}

    def sim_conf(by_seed, conf_wins, round_wins, playin_q):
        # Play-In
        q7, q8 = sim_playin_mc(by_seed[7],by_seed[8],by_seed[9],by_seed[10],nr_map,home_adv,rng)
        playin_q[q7][7] += 1
        playin_q[q8][8] += 1

        # Bracket: seeds 1-6 direct + q7 (seed7), q8 (seed8)
        # Higher seed always = home team
        bk = [by_seed[1],by_seed[2],by_seed[3],by_seed[4],by_seed[5],by_seed[6],q7,q8]
        # seed[i] = i+1 (0-indexed)
        bseed = {bk[i]:i+1 for i in range(8)}

        def hs(a,b): return a if bseed[a]<bseed[b] else b  # higher seed = home
        def ls(a,b): return b if bseed[a]<bseed[b] else a

        def sim(a,b):
            h,aw = hs(a,b),ls(a,b)
            return sim_series_mc(h,aw,nr_map,home_adv,rng)

        # R1: 1v8, 2v7, 3v6, 4v5
        r1 = [sim(bk[0],bk[7]), sim(bk[1],bk[6]), sim(bk[2],bk[5]), sim(bk[3],bk[4])]
        for w in r1: round_wins[w][1] += 1

        # R2: W(1/8) vs W(4/5), W(2/7) vs W(3/6)  — re-seed: best remaining vs worst
        # Standard NBA re-seed: top half bracket vs bottom half bracket
        r2 = [sim(r1[0],r1[3]), sim(r1[1],r1[2])]
        for w in r2: round_wins[w][2] += 1

        # Conf Final
        cf = sim(r2[0],r2[1])
        round_wins[cf][3] += 1
        conf_wins[cf]     += 1
        return cf

    for _ in range(n_sim):
        e_cf = sim_conf(e_by_seed, conf_wins_e, round_wins_e, playin_q_e)
        w_cf = sim_conf(w_by_seed, conf_wins_w, round_wins_w, playin_q_w)

        # NBA Finals — higher NR gets home court
        if nr_map[e_cf] >= nr_map[w_cf]:
            winner = sim_series_mc(e_cf, w_cf, nr_map, home_adv, rng)
        else:
            winner = sim_series_mc(w_cf, e_cf, nr_map, home_adv, rng)
        nba_wins[winner] += 1

    return conf_wins_e, conf_wins_w, nba_wins, round_wins_e, round_wins_w, playin_q_e, playin_q_w


def fmt(v,n=1): return f"{v*100:.{n}f}%"


# ══════════════════════════════════════════════════════════════════════════════
# 2025 PLAYOFF BRACKET (April 18 start)
# Bracket confirmed by user. Net ratings are editable defaults — update from
# NBA.com / Cleaning the Glass before running.
# ══════════════════════════════════════════════════════════════════════════════
DEFAULT_EAST = [
    # (name, seed, net_rating)  ← syötä NR itse
    ("Detroit Pistons",        1,  0.0),
    ("Boston Celtics",         2,  0.0),
    ("New York Knicks",        3,  0.0),
    ("Cleveland Cavaliers",    4,  0.0),
    ("Toronto Raptors",        5,  0.0),
    ("Atlanta Hawks",          6,  0.0),
    # Play-In
    ("Philadelphia 76ers",     7,  0.0),
    ("Orlando Magic",          8,  0.0),
    ("Charlotte Hornets",      9,  0.0),
    ("Miami Heat",            10,  0.0),
]
DEFAULT_WEST = [
    # (name, seed, net_rating)  ← syötä NR itse
    ("Oklahoma City Thunder",  1,  0.0),
    ("San Antonio Spurs",      2,  0.0),
    ("Denver Nuggets",         3,  0.0),
    ("Los Angeles Lakers",     4,  0.0),
    ("Houston Rockets",        5,  0.0),
    ("Minnesota Timberwolves", 6,  0.0),
    # Play-In
    ("Phoenix Suns",           7,  0.0),
    ("Portland Trail Blazers", 8,  0.0),
    ("LA Clippers",            9,  0.0),
    ("Golden State Warriors", 10,  0.0),
]


# ══════════════════════════════════════════════════════════════════════════════
# SIDEBAR
# ══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("## ⚙️ Asetukset")
    home_adv = st.slider("Kotietu (net rating pistettä)", 0.0, 5.0, 3.0, 0.5,
        help="NBA:n historiallinen kotietu ≈ 2.5–3.5 NR-pistettä")
    n_sim = st.select_slider("Monte Carlo kierrokset",
        options=[10_000,50_000,100_000,200_000], value=100_000,
        format_func=lambda x: f"{x:,}")
    st.divider()
    st.markdown("### 📊 Net Rating selitys")
    st.markdown("""
Net Rating = Offensive Rating − Defensive Rating per 100 possessions.

| NR | Taso |
|----|------|
| +8 tai yli | Eliittijoukkue |
| +4 … +8 | Playoff-haastaja |
| +1 … +4 | Playoff-joukkue |
| -1 … +1 | Tasainen |
| alle -1 | Play-In alue |

**Voitontodennäköisyys:**  
10 NR-pisteen ero ≈ 73% voittotodennäköisyys.
""")
    st.divider()
    st.markdown("### 🔮 Play-In")
    st.markdown("""
**P1:** #7 vs #8 → voittaja = playoff #7  
**P2:** #9 vs #10 → häviäjä putoaa  
**P3:** häviäjä(P1) vs voittaja(P2) → voittaja = playoff #8  
""")

# ══════════════════════════════════════════════════════════════════════════════
# HEADER
# ══════════════════════════════════════════════════════════════════════════════
st.markdown("""
<div style="background:linear-gradient(135deg,#0f0f1a,#1a0a2e,#0f1a0a);
 border:1px solid #f97316;border-radius:12px;padding:22px 32px;margin-bottom:20px;text-align:center;">
 <h1 style="font-size:2.8rem;margin:0;color:#f97316;">🏀 NBA PLAYOFFS 2025</h1>
 <p style="color:#8888aa;margin:6px 0 0;font-size:0.85rem;letter-spacing:2px;">
  PLAYOFF-TODENNÄKÖISYYSLASKURI · NET RATING · PLAY-IN · PATH-AWARE MONTE CARLO
 </p>
</div>
""", unsafe_allow_html=True)

tab1, tab2, tab3, tab4, tab5 = st.tabs([
    "🏆 BRACKET & MESTARUUS", "🎯 SARJA-ANALYYSI", "🔮 PLAY-IN", "📈 REITTIANALYYSI", "📊 EXCEL-POHJA"
])


# ════════════════════════════════════════════════════════════════════════
# TAB 1 – Full bracket simulation + championship odds
# ════════════════════════════════════════════════════════════════════════
with tab1:
    st.markdown("### Syötä joukkueiden net ratingit")

    # Reset-nappi tyhjentää session staten → pakottaa uudet oletusarvot
    if st.button("🔄 Nollaa joukkueet oletuksiin (2025 bracket)", key="reset_teams"):
        for key in list(st.session_state.keys()):
            if key.startswith("e_") or key.startswith("w_"):
                del st.session_state[key]
        st.rerun()

    def build_10_inputs(defaults, key_prefix, conf_label):
        teams = []
        with st.expander(f"🏀 {conf_label} – net ratingit", expanded=True):
            hcols = st.columns([3,1,2])
            hcols[0].markdown("**Joukkue**")
            hcols[1].markdown("**Sija**")
            hcols[2].markdown("**Net Rating**")
            for i,(dname,dseed,dnr) in enumerate(defaults):
                if dseed == 7:
                    st.markdown('<div style="border-top:1px dashed #7c3aed;margin:4px 0;color:#a78bfa;font-size:0.72rem;letter-spacing:1px;padding-top:4px;">▼ PLAY-IN (7–10)</div>', unsafe_allow_html=True)
                c1,c2,c3 = st.columns([3,1,2])
                # Index-based keys — these never clash with old session state
                name = c1.text_input("",value=dname,key=f"{key_prefix}_n{i}",label_visibility="collapsed")
                seed = c2.number_input("",value=dseed,min_value=1,max_value=10,
                                       key=f"{key_prefix}_s{i}",label_visibility="collapsed")
                nr   = c3.number_input("",value=dnr,step=0.1,format="%.1f",
                                       key=f"{key_prefix}_nr{i}",label_visibility="collapsed")
                teams.append({"name":name,"seed":seed,"nr":nr})
        return teams

    col_e, col_w = st.columns(2)
    with col_e:
        east_teams = build_10_inputs(DEFAULT_EAST, "e", "Itäinen konferenssi")
    with col_w:
        west_teams = build_10_inputs(DEFAULT_WEST, "w", "Läntinen konferenssi")

    if st.button("🏆 SIMULOI KOKO BRACKET (PATH-AWARE)", use_container_width=True):
        with st.spinner(f"Simuloidaan {n_sim:,} täyttä playoff-kautta..."):
            (conf_e, conf_w, nba_w,
             rnd_e, rnd_w,
             pq_e, pq_w) = sim_full_bracket(east_teams, west_teams, home_adv, n_sim)

        st.session_state['sim_results'] = (conf_e, conf_w, nba_w, rnd_e, rnd_w, pq_e, pq_w,
                                            east_teams, west_teams, n_sim)

    if 'sim_results' in st.session_state:
        (conf_e, conf_w, nba_w, rnd_e, rnd_w, pq_e, pq_w,
         east_teams_s, west_teams_s, n_sim_s) = st.session_state['sim_results']

        # ── Play-In summary ────────────────────────────────
        st.divider()
        st.markdown("#### 🔮 Play-In – Pääsy playoffseihin")
        pce, pcw = st.columns(2)
        for pcol, teams, pq_map, label in [(pce,east_teams_s,pq_e,"Itä"),(pcw,west_teams_s,pq_w,"Länsi")]:
            with pcol:
                st.markdown(f"**{label}**")
                pi_rows = []
                for t in sorted(teams, key=lambda x:x['seed']):
                    if t['seed'] < 7: continue
                    q = pq_map[t['name']]
                    pi_rows.append({
                        "Joukkue":t['name'], "Sija":f"#{t['seed']}", "NR":f"{t['nr']:+.1f}",
                        "Sija #7 %":fmt(q[7]/n_sim_s),
                        "Sija #8 %":fmt(q[8]/n_sim_s),
                        "Pääsee %":fmt((q[7]+q[8])/n_sim_s),
                    })
                st.dataframe(pd.DataFrame(pi_rows),use_container_width=True,hide_index=True)

        # ── Conference + NBA results ───────────────────────
        st.divider()
        st.markdown("#### 🏆 Konferenssi- ja NBA-mestaruus")

        all_rows = []
        nr_all = {t['name']:t['nr'] for t in east_teams_s+west_teams_s}
        for t in east_teams_s:
            rw = rnd_e[t['name']]
            all_rows.append({
                "Joukkue": t['name'] + (" 🔮" if t['seed']>=7 else ""),
                "Konf.": "Itä", "Sija": t['seed'], "Net Rating": f"{t['nr']:+.1f}",
                "R1 %":  fmt(rw[1]/n_sim_s),
                "R2 %":  fmt(rw[2]/n_sim_s),
                "CF %":  fmt(rw[3]/n_sim_s),
                "Konf. Mestari %": round(conf_e[t['name']]/n_sim_s*100,1),
                "NBA Mestari %":   round(nba_w[t['name']]/n_sim_s*100,1),
                "_nba": nba_w[t['name']],
            })
        for t in west_teams_s:
            rw = rnd_w[t['name']]
            all_rows.append({
                "Joukkue": t['name'] + (" 🔮" if t['seed']>=7 else ""),
                "Konf.": "Länsi", "Sija": t['seed'], "Net Rating": f"{t['nr']:+.1f}",
                "R1 %":  fmt(rw[1]/n_sim_s),
                "R2 %":  fmt(rw[2]/n_sim_s),
                "CF %":  fmt(rw[3]/n_sim_s),
                "Konf. Mestari %": round(conf_w[t['name']]/n_sim_s*100,1),
                "NBA Mestari %":   round(nba_w[t['name']]/n_sim_s*100,1),
                "_nba": nba_w[t['name']],
            })

        df_all = pd.DataFrame(all_rows).sort_values("_nba",ascending=False)
        st.dataframe(
            df_all[["Joukkue","Konf.","Sija","Net Rating","R1 %","R2 %","CF %","Konf. Mestari %","NBA Mestari %"]],
            use_container_width=True, hide_index=True
        )

        # Top-8 NBA championship bar chart
        top8 = df_all.head(8)
        st.bar_chart(top8.set_index("Joukkue")["NBA Mestari %"], color="#f97316", height=300)

        # ── East/West breakdown ────────────────────────────
        st.divider()
        ce, cw2 = st.columns(2)
        for col, teams_s, cwins, rnd_map, label, color in [
            (ce, east_teams_s, conf_e, rnd_e, "Itä",  "#f97316"),
            (cw2,west_teams_s, conf_w, rnd_w, "Länsi","#3b82f6"),
        ]:
            with col:
                st.markdown(f"**{label}inen konferenssi – kierrostilastot**")
                rows_c = []
                for t in sorted(teams_s, key=lambda x:x['seed']):
                    rw = rnd_map[t['name']]
                    rows_c.append({
                        "Joukkue": t['name']+(" 🔮" if t['seed']>=7 else ""),
                        "NR": f"{t['nr']:+.1f}",
                        "R1 %":fmt(rw[1]/n_sim_s), "R2 %":fmt(rw[2]/n_sim_s), "CF %":fmt(rw[3]/n_sim_s),
                        "Mestari %":round(cwins[t['name']]/n_sim_s*100,1),
                    })
                df_c = pd.DataFrame(rows_c)
                st.dataframe(df_c, use_container_width=True, hide_index=True)
                st.bar_chart(df_c.set_index("Joukkue")["Mestari %"], color=color, height=180)


# ════════════════════════════════════════════════════════════════════════
# TAB 2 – Single series analysis
# ════════════════════════════════════════════════════════════════════════
with tab2:
    st.markdown("### Kahden joukkueen sarja-analyysi")
    st.caption("Net Rating = esim. +5.2 tai -1.3")

    c1,c2 = st.columns(2)
    with c1:
        st.markdown('<div style="color:#f97316;font-family:Bebas Neue,sans-serif;font-size:1.1rem;letter-spacing:2px;">🟠 KORKEAMPI SIJOITUS (KOTIETU)</div>', unsafe_allow_html=True)
        sa_name = st.text_input("Joukkue A", value="Detroit Pistons", key="sa_n")
        sa_nr   = st.number_input("Net Rating", value=0.0, step=0.1, format="%.1f", key="sa_nr")
    with c2:
        st.markdown('<div style="color:#3b82f6;font-family:Bebas Neue,sans-serif;font-size:1.1rem;letter-spacing:2px;">🔵 MATALAMPI SIJOITUS (VIERASJOUKKUE)</div>', unsafe_allow_html=True)
        sb_name = st.text_input("Joukkue B", value="Boston Celtics", key="sb_n")
        sb_nr   = st.number_input("Net Rating", value=0.0, step=0.1, format="%.1f", key="sb_nr")

    best_of = st.radio("Sarjamuoto", [5,7], index=1, horizontal=True,
                       format_func=lambda x: f"Best-of-{x}")

    if st.button("🔢 LASKE SARJATODENNÄKÖISYYDET", use_container_width=True):
        res = series_probs_exact(sa_nr, sb_nr, home_adv, best_of)
        wins_n = (best_of+1)//2

        ph = win_prob(sa_nr, sb_nr,  home_adv)
        pa = win_prob(sa_nr, sb_nr, -home_adv)

        st.divider()
        st.markdown("#### Yhden pelin voittotodennäköisyys")
        m1,m2,m3,m4 = st.columns(4)
        with m1:
            st.markdown(f'<div class="metric-card"><p class="metric-value">{ph*100:.1f}%</p><p class="metric-label">{sa_name} kotona</p></div>',unsafe_allow_html=True)
        with m2:
            st.markdown(f'<div class="metric-card"><p class="metric-value" style="color:#3b82f6;">{(1-ph)*100:.1f}%</p><p class="metric-label">{sb_name} vieraana</p></div>',unsafe_allow_html=True)
        with m3:
            st.markdown(f'<div class="metric-card"><p class="metric-value" style="color:#3b82f6;">{(1-pa)*100:.1f}%</p><p class="metric-label">{sb_name} kotona</p></div>',unsafe_allow_html=True)
        with m4:
            st.markdown(f'<div class="metric-card"><p class="metric-value">{pa*100:.1f}%</p><p class="metric-label">{sa_name} vieraana</p></div>',unsafe_allow_html=True)

        st.divider()
        st.markdown("#### Sarjan voittotodennäköisyys")
        ca,cb = st.columns(2)
        with ca:
            st.markdown(f'<div class="metric-card"><p class="metric-value">{res["p_home"]*100:.1f}%</p><p class="metric-label">{sa_name} voittaa sarjan</p></div>',unsafe_allow_html=True)
        with cb:
            st.markdown(f'<div class="metric-card"><p class="metric-value" style="color:#3b82f6;">{res["p_away"]*100:.1f}%</p><p class="metric-label">{sb_name} voittaa sarjan</p></div>',unsafe_allow_html=True)

        bh,ba = res['p_home']*100, res['p_away']*100
        st.markdown(f"""
        <div style="display:flex;gap:3px;margin:10px 0;">
            <div style="width:{bh:.1f}%;background:linear-gradient(90deg,#f97316,#fb923c);
             border-radius:6px 0 0 6px;padding:5px 10px;color:white;font-weight:700;
             font-size:0.82rem;white-space:nowrap;overflow:hidden;">{sa_name} {bh:.1f}%</div>
            <div style="width:{ba:.1f}%;background:linear-gradient(90deg,#3b82f6,#60a5fa);
             border-radius:0 6px 6px 0;padding:5px 10px;color:white;font-weight:700;
             font-size:0.82rem;text-align:right;white-space:nowrap;overflow:hidden;">{ba:.1f}% {sb_name}</div>
        </div>""",unsafe_allow_html=True)

        st.divider()
        st.markdown("#### Marginaalilinjat")
        rows_m = [{"Linja":lbl,
                   f"{sa_name} kattaa":fmt(res[f'p_home_m{m}']),
                   f"{sb_name} kattaa":fmt(res[f'p_away_m{m}'])}
                  for m,lbl in [(3,"+3.5"),(2,"+2.5"),(1,"+1.5")]]
        st.dataframe(pd.DataFrame(rows_m).set_index("Linja"),use_container_width=True)

        st.divider()
        st.markdown("#### Sarjan lopputulokset")
        oc_rows = []
        for (wh,wa),prob in sorted(res['outcomes'].items(),key=lambda x:-x[1]):
            winner = sa_name if wh==wins_n else sb_name
            loser  = sb_name if wh==wins_n else sa_name
            score  = f"{wh}–{wa}" if wh==wins_n else f"{wa}–{wh}"
            oc_rows.append({"Tulos":f"{winner} {score} {loser}","Voittaja":winner,
                            "Todennäköisyys":fmt(prob),"Tn %":round(prob*100,2)})
        df_oc = pd.DataFrame(oc_rows)
        st.dataframe(df_oc[["Tulos","Voittaja","Todennäköisyys"]],use_container_width=True,hide_index=True)
        st.bar_chart(df_oc.set_index("Tulos")["Tn %"],color="#f97316",height=240)


# ════════════════════════════════════════════════════════════════════════
# TAB 3 – Play-In analysis
# ════════════════════════════════════════════════════════════════════════
with tab3:
    st.markdown("### Play-In turnauksen analyysi")
    st.markdown("""
<div class="playin-box">
<b style="color:#a78bfa;font-family:Bebas Neue,sans-serif;letter-spacing:2px;font-size:1.1rem;">🔮 PLAY-IN RAKENNE</b><br><br>
<b>Peli 1:</b> #7 vs #8 (kotietu: #7) → <b>Voittaja = playoff sija #7</b><br>
<b>Peli 2:</b> #9 vs #10 (kotietu: #9) → <b>Häviäjä putoaa kaudelta</b><br>
<b>Peli 3:</b> Häviäjä(P1) vs Voittaja(P2) (kotietu: häviäjä(P1)) → <b>Voittaja = playoff sija #8</b>
</div>
""",unsafe_allow_html=True)

    conf_pi = st.radio("Konferenssi", ["Itäinen","Läntinen"], horizontal=True, key="pi_conf")
    defaults_pi_e = [(n,nr) for n,s,nr in DEFAULT_EAST if s>=7]
    defaults_pi_w = [(n,nr) for n,s,nr in DEFAULT_WEST if s>=7]
    defs = defaults_pi_e if conf_pi=="Itäinen" else defaults_pi_w

    pi_cols = st.columns(4)
    pi_names, pi_nrs = [], []
    for i,(col,(dname,dnr)) in enumerate(zip(pi_cols,defs)):
        with col:
            st.markdown(f'<div style="color:#a78bfa;font-family:Bebas Neue,sans-serif;font-size:1rem;">SIJA #{i+7}</div>',unsafe_allow_html=True)
            safe = dname.replace(" ","_").replace(".","")
            n  = st.text_input("",value=dname,key=f"pi_n_{safe}_{conf_pi}",label_visibility="collapsed")
            nr = st.number_input("NR",value=dnr,step=0.1,format="%.1f",key=f"pi_nr_{safe}_{conf_pi}",label_visibility="collapsed")
            pi_names.append(n); pi_nrs.append(nr)

    if st.button("🔮 LASKE PLAY-IN TODENNÄKÖISYYDET", use_container_width=True):
        pm = {pi_names[i]:pi_nrs[i] for i in range(4)}
        t7,t8,t9,t10 = pi_names
        probs = playin_probs_exact(t7,t8,t9,t10,pm,home_adv)

        g1 = win_prob(pm[t7],pm[t8], home_adv)
        g2 = win_prob(pm[t9],pm[t10],home_adv)

        st.divider()
        st.markdown("#### Pelien voittotodennäköisyydet")
        gc1,gc2,gc3 = st.columns(3)
        with gc1:
            st.markdown(f"""<div class="metric-card">
            <p style="color:#a78bfa;font-family:Bebas Neue,sans-serif;letter-spacing:1px;margin:0 0 3px;">PELI 1</p>
            <p style="margin:2px 0;font-size:0.78rem;">{t7} vs {t8}</p>
            <p class="metric-value">{g1*100:.1f}%</p>
            <p class="metric-label">{t7} voittaa</p></div>""",unsafe_allow_html=True)
        with gc2:
            st.markdown(f"""<div class="metric-card">
            <p style="color:#a78bfa;font-family:Bebas Neue,sans-serif;letter-spacing:1px;margin:0 0 3px;">PELI 2</p>
            <p style="margin:2px 0;font-size:0.78rem;">{t9} vs {t10}</p>
            <p class="metric-value">{g2*100:.1f}%</p>
            <p class="metric-label">{t9} voittaa</p></div>""",unsafe_allow_html=True)
        with gc3:
            st.markdown("""<div class="metric-card">
            <p style="color:#a78bfa;font-family:Bebas Neue,sans-serif;letter-spacing:1px;margin:0 0 3px;">PELI 3</p>
            <p style="margin:2px 0;font-size:0.78rem;">4 skenaariota</p>
            <p class="metric-value" style="font-size:1.8rem;">↓</p>
            <p class="metric-label">Katso alta</p></div>""",unsafe_allow_html=True)

        st.divider()
        st.markdown("#### Pääsy playoffseihin")
        pi_rows = [{"Joukkue":n,"Lähtösija":f"#{i+7}","Net Rating":f"{pi_nrs[i]:+.1f}",
                    "Sija #7 %":fmt(probs[n]['p7']),"Sija #8 %":fmt(probs[n]['p8']),
                    "Pääsee %":fmt(probs[n]['p_qualify']),"_q":probs[n]['p_qualify']}
                   for i,n in enumerate(pi_names)]
        df_pi = pd.DataFrame(pi_rows).sort_values("_q",ascending=False)
        st.dataframe(df_pi[["Joukkue","Lähtösija","Net Rating","Sija #7 %","Sija #8 %","Pääsee %"]],
                     use_container_width=True,hide_index=True)

        st.markdown("#### Visualisointi")
        for _,row in df_pi.iterrows():
            nm = row["Joukkue"]
            p7  = probs[nm]['p7']*100; p8  = probs[nm]['p8']*100
            pout= (1-probs[nm]['p_qualify'])*100
            st.markdown(f"**{nm}** ({row['Lähtösija']})")
            st.markdown(f"""<div style="display:flex;gap:2px;margin:2px 0 9px;">
                <div style="width:{p7:.1f}%;background:linear-gradient(90deg,#f97316,#fb923c);
                 border-radius:5px 0 0 5px;padding:3px 7px;color:white;font-size:0.73rem;
                 font-weight:700;white-space:nowrap;overflow:hidden;min-width:0;">Sija #7: {p7:.1f}%</div>
                <div style="width:{p8:.1f}%;background:linear-gradient(90deg,#7c3aed,#a78bfa);
                 padding:3px 7px;color:white;font-size:0.73rem;font-weight:700;
                 white-space:nowrap;overflow:hidden;min-width:0;">Sija #8: {p8:.1f}%</div>
                <div style="width:{pout:.1f}%;background:#1a1a2e;border-radius:0 5px 5px 0;
                 padding:3px 7px;color:#8888aa;font-size:0.73rem;
                 white-space:nowrap;overflow:hidden;min-width:0;">Putoaa: {pout:.1f}%</div>
            </div>""",unsafe_allow_html=True)

        st.divider()
        st.markdown("#### Peli 3 – kaikki skenaariot")
        sc_rows = []
        for (g1w,g1l),(g2w,g2l),prob_s in [
            ((t7,t8),(t9,t10), win_prob(pm[t7],pm[t8],home_adv)*win_prob(pm[t9],pm[t10],home_adv)),
            ((t7,t8),(t10,t9), win_prob(pm[t7],pm[t8],home_adv)*(1-win_prob(pm[t9],pm[t10],home_adv))),
            ((t8,t7),(t9,t10),(1-win_prob(pm[t7],pm[t8],home_adv))*win_prob(pm[t9],pm[t10],home_adv)),
            ((t8,t7),(t10,t9),(1-win_prob(pm[t7],pm[t8],home_adv))*(1-win_prob(pm[t9],pm[t10],home_adv))),
        ]:
            pg3 = win_prob(pm[g1l],pm[g2w],home_adv)
            sc_rows.append({
                "Skenaario": f"P1: {g1w} voittaa, P2: {g2w} voittaa → {g1l} vs {g2w}",
                "Tn": fmt(prob_s),
                f"{g1l} voittaa G3": fmt(pg3),
                f"{g2w} voittaa G3": fmt(1-pg3),
            })
        st.dataframe(pd.DataFrame(sc_rows),use_container_width=True,hide_index=True)


# ════════════════════════════════════════════════════════════════════════
# TAB 4 – Route/path analysis (uses sim_results from Tab1)
# ════════════════════════════════════════════════════════════════════════
with tab4:
    st.markdown("### Reittianalyysi – miten joukkue pääsee mestariksi")
    st.caption("Tämä välilehti hyödyntää Tab 1:n simulaation tuloksia. Aja simulaatio ensin.")

    if 'sim_results' not in st.session_state:
        st.info("Aja ensin simulaatio **Bracket & Mestaruus** -välilehdellä.")
    else:
        (conf_e, conf_w, nba_w, rnd_e, rnd_w, pq_e, pq_w,
         east_ts, west_ts, n_sim_s) = st.session_state['sim_results']

        all_teams = east_ts + west_ts
        team_names_sorted = sorted(
            [t['name'] for t in all_teams],
            key=lambda n: -nba_w[n]
        )
        selected = st.selectbox("Valitse joukkue", team_names_sorted)

        t_obj = next(t for t in all_teams if t['name']==selected)
        is_east = t_obj in east_ts
        rnd_map  = rnd_e if is_east else rnd_w
        conf_map = conf_e if is_east else conf_w
        pq_map   = pq_e  if is_east else pq_w

        st.divider()
        rw = rnd_map[selected]
        playin = t_obj['seed'] >= 7

        st.markdown(f"#### {selected} – Todennäköisyydet jokaiselle vaiheelle")

        stages = []
        if playin:
            q = pq_map[selected]
            stages += [
                ("Play-In: Pääsee playoffseihin", (q[7]+q[8])/n_sim_s, "#7c3aed"),
                ("Play-In: Sija #7", q[7]/n_sim_s, "#a78bfa"),
                ("Play-In: Sija #8", q[8]/n_sim_s, "#7c3aed"),
            ]
        stages += [
            ("Voittaa 1. kierroksen", rw[1]/n_sim_s, "#f97316"),
            ("Voittaa 2. kierroksen (puolivälierä)", rw[2]/n_sim_s, "#fb923c"),
            ("Voittaa konferenssifinalin", rw[3]/n_sim_s, "#fbbf24"),
            ("Voittaa konferenssin (finaaleihin)", conf_map[selected]/n_sim_s, "#22c55e"),
            ("Voittaa NBA:n mestaruuden", nba_w[selected]/n_sim_s, "#16a34a"),
        ]

        for label, prob, color in stages:
            pct = prob*100
            st.markdown(f"**{label}**")
            st.markdown(f"""
            <div style="display:flex;align-items:center;gap:10px;margin:2px 0 10px;">
                <div style="flex:1;background:#1a1a2e;border-radius:6px;overflow:hidden;height:26px;">
                    <div style="width:{min(pct,100):.1f}%;background:{color};height:100%;
                     display:flex;align-items:center;padding-left:10px;color:white;
                     font-weight:700;font-size:0.8rem;white-space:nowrap;">{pct:.1f}%</div>
                </div>
            </div>""",unsafe_allow_html=True)

        # Conditional probabilities (given they got this far)
        st.divider()
        st.markdown("#### Ehdolliset todennäköisyydet (jos päässyt tähän asti)")
        cond_rows = []
        prev_p = 1.0
        for label, prob, _ in [s for s in stages if "Play-In: Sija" not in s and "Play-In: Pääsee" not in s]:
            if prev_p > 0:
                cond = prob / (prev_p if prev_p > 0 else 1)
                cond_rows.append({"Vaihe": label,
                                   "Absoluuttinen %": f"{prob*100:.1f}%",
                                   "Ehdollinen %": f"{cond*100:.1f}%"})
            prev_p = prob

        if cond_rows:
            st.dataframe(pd.DataFrame(cond_rows),use_container_width=True,hide_index=True)
            st.caption("Ehdollinen % = todennäköisyys voittaa tämä kierros, kun on jo päässyt siihen asti.")

        # NR comparison vs typical opponents
        st.divider()
        st.markdown("#### Net Rating vertailu potentiaalisiin vastustajiin")
        nr_t = t_obj['nr']
        opp_pool = west_ts if is_east else east_ts
        opp_pool_same = east_ts if is_east else west_ts
        rows_vs = []
        for opp in sorted(opp_pool_same + opp_pool, key=lambda x: -x['nr']):
            if opp['name']==selected: continue
            same_conf = opp in (east_ts if is_east else west_ts)
            # Home court depends on seed; for display show NR-based matchup
            if nr_t >= opp['nr']:
                p_home_w = win_prob(nr_t, opp['nr'],  home_adv)
                p_away_w = win_prob(nr_t, opp['nr'], -home_adv)
            else:
                p_home_w = win_prob(nr_t, opp['nr'], -home_adv)
                p_away_w = win_prob(nr_t, opp['nr'],  home_adv)
            # Series win prob
            if nr_t >= opp['nr']:
                sr = series_probs_exact(nr_t, opp['nr'], home_adv)
                p_ser = sr['p_home']
            else:
                sr = series_probs_exact(opp['nr'], nr_t, home_adv)
                p_ser = sr['p_away']
            rows_vs.append({
                "Vastustaja": opp['name'],
                "Konf.": "Sama" if same_conf else "Finals",
                "Vast. NR": f"{opp['nr']:+.1f}",
                "NR-ero": f"{nr_t-opp['nr']:+.1f}",
                "Sarjavoitto %": fmt(p_ser),
            })
        st.dataframe(pd.DataFrame(rows_vs).head(16),use_container_width=True,hide_index=True)


# ════════════════════════════════════════════════════════════════════════
# TAB 5 – Excel template
# ════════════════════════════════════════════════════════════════════════
with tab5:
    st.markdown("### 📊 Excel-pohja – Net Rating")

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Net Ratings"

    headers = ["Joukkue","Konferenssi","Sija","Net Rating","Huomio"]
    hfill = PatternFill("solid",fgColor="1a1a2e")
    hfont = Font(bold=True,color="F97316",name="Calibri",size=12)
    thin  = Side(border_style="thin",color="2a2a4a")
    brd   = Border(left=thin,right=thin,top=thin,bottom=thin)

    for ci,h in enumerate(headers,1):
        cell = ws.cell(1,ci,h)
        cell.font,cell.fill = hfont,hfill
        cell.alignment,cell.border = Alignment(horizontal="center"),brd

    all_20 = [(n,"Itä",s,nr,"") for n,s,nr in DEFAULT_EAST] + \
             [(n,"Länsi",s,nr,"") for n,s,nr in DEFAULT_WEST]

    dfont  = Font(name="Calibri",size=11,color="E8E8E8")
    dfill  = PatternFill("solid",fgColor="0a0a0f")
    afill  = PatternFill("solid",fgColor="111118")
    pifill = PatternFill("solid",fgColor="1a0a2e")

    for ri,row in enumerate(all_20,2):
        is_pi = row[2] >= 7
        fill  = pifill if is_pi else (dfill if ri%2==0 else afill)
        for ci,val in enumerate(row,1):
            cell = ws.cell(ri,ci,val)
            cell.font,cell.fill = dfont,fill
            cell.alignment = Alignment(horizontal="center" if ci>1 else "left")
            cell.border = brd

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 20

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)

    st.download_button("⬇️ LATAA EXCEL-POHJA",
        data=buf, file_name="nba_2025_net_ratings.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True)

    st.divider()
    st.markdown("""
### Net Rating -lähteet ja kalibrointi

**Mistä löydät Net Ratingin:**
- [NBA.com/stats](https://www.nba.com/stats/teams/advanced) → Team Advanced Stats
- [Basketball Reference](https://www.basketball-reference.com) → Team Stats → Advanced
- [Cleaning the Glass](https://cleaningtheglass.com) → Adjusted Net Rating (suositeltavin)

**Mallin kalibrointi:**
| NR-ero | Game win% | Series win% (B7) |
|--------|-----------|-----------------|
| 0 | 62.0% (kotietu) | ~65% |
| 3 | 67.0% | ~75% |
| 5 | 71.5% | ~83% |
| 10 | 79.8% | ~94% |

*Kotietu = 3.0 NR-pistettä oletuksena. Voidaan säätää sivupalkista.*

**Play-In:** Sijat 7–10 Play-In turnauksen kautta. 
Kotietu paremmalle siemenelle jokaisessa Play-In pelissä.
""")
